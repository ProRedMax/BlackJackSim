import openpyxl

import Card
import Deck

bj_stat = {}  # Player Value : dict(Dealer Up Card, Tuple(GamesPlayedWithHand, Won, Lost)
bj_remember = {}  # Player Value : dict(Dealer Up Card : Action/Hit)
verbose = False


def calc_value(cards):
    value = 0
    num_aces = 0

    for card in cards:
        if card.isAce:
            num_aces += 1
        value += card.value

    # Aces
    while num_aces > 0 and value > 21:
        value -= 10  # Reduce the value from the ace from 11 to 1
        num_aces -= 1

    return value


def remember_round(my_value, dealer_value, gamestate, var: {}):
    """
    Adds match to the statistic
    :param var: Storage
    :param my_value: Value on players hand
    :param dealer_value: Dealers value after end of game
    :param gamestate: 0... Push, 1... Loss, 2... Won
    """
    stat = var
    if my_value in stat:
        temp_map = stat[my_value]
        if dealer_value in temp_map:
            val_stat = temp_map[dealer_value]
            games_played = val_stat[0]
            won = val_stat[1]
            lost = val_stat[2]
            if gamestate == 1:
                lost = lost + 1
            elif gamestate == 2:
                won = won + 1
            games_played = games_played + 1
            temp_map[dealer_value] = (games_played, won, lost)
        else:
            loose = 0
            win = 0
            if gamestate == 1:
                loose = loose + 1
            elif gamestate == 2:
                win = win + 1
            temp_map[dealer_value] = (1, win, loose)
        stat[my_value] = temp_map
    else:
        loose = 0
        win = 0
        if gamestate == 1:
            loose = loose + 1
        elif gamestate == 2:
            win = win + 1
        stat[my_value] = {dealer_value: (1, win, loose)}


def trial_error_game(num_of_players: int = 1):
    current_deck = Deck.Deck(8)
    current_deck.shuffle()
    my_cards = []
    dealer_upcard: Card.Card
    dealer_cards = []

    # Deal
    my_cards = [current_deck.pop()]
    dealer_upcard = current_deck.pop()
    dealer_cards = [dealer_upcard]
    my_cards.append(current_deck.pop())

    if verbose: print(("My cards: " + str(my_cards)))
    if verbose: print(("Dealer cards:", dealer_cards))

    value = calc_value(my_cards)
    if verbose: print((value))

    if value == 21:  # BlackJack
        remember_round(value, dealer_upcard.value, 2)
        return

    while value in bj_remember:
        actions = bj_remember[value]
        if dealer_upcard.value in actions:
            action = actions[dealer_upcard.value]
            if action:  # Hit
                if verbose: print(("Hitting with " + str(value) + " on a dealers " + str(dealer_upcard.value)))
                pulled_card = current_deck.pop()
                if verbose: print(("We pulled a " + str(pulled_card.value)))
                my_cards.append(pulled_card)
                value = calc_value(my_cards)
                if value > 21:  # Bust
                    if verbose: print(("We busted!"))
                    remember_round(value, dealer_upcard.value, 1, bj_stat)
                    actions[dealer_upcard.value] = False
                    bj_remember[value] = actions
                    return
            else:
                break
        else:
            break

    value = calc_value(my_cards)
    if verbose: print(("Standing on " + str(value)))
    # Dealer gets cards
    if verbose: print(("Dealer has " + str(dealer_upcard.value)))
    while calc_value(dealer_cards) < 17:
        new_card = current_deck.pop()
        if verbose: print(("Dealer pulled " + str(new_card.value)))
        dealer_cards.append(new_card)
        if verbose: print(("Dealer has " + str(calc_value(dealer_cards))))

    if calc_value(dealer_cards) > 21:  # Dealer busted
        if verbose: print(("Dealer busted!"))
        dealer_cards = [Card.Card(0, False)]

    if calc_value(dealer_cards) > value:  # Loss
        remember_round(value, dealer_upcard.value, 1, bj_stat)
        if verbose: print(("Loss"))
        # Remember move for next round
        if value in bj_remember:
            actions = bj_remember[value]
            if dealer_upcard.value in actions:
                action = actions[dealer_upcard.value]
                actions[dealer_upcard.value] = not action
            else:
                actions[dealer_upcard.value] = True
        else:
            bj_remember[value] = {dealer_upcard.value: True}

    elif calc_value(dealer_cards) < value:  # Win
        remember_round(value, dealer_upcard.value, 2, bj_stat)
        if verbose: print(("Win"))
    else:  # Push
        remember_round(value, dealer_upcard.value, 0, bj_stat)
        if verbose: print(("Push"))

    # End of round


bj_bs_hit_table = {}  # Player up card : dict:DealerUpCard : (GamesPlayed, Win, Loss)
bj_bs_stay_table = {}  # Player up card : dict:DealerUpCard : (GamesPlayed, Win, Loss)


def stat_game(current_deck: Deck.Deck, stay: bool = False):
    if len(current_deck.cards) < 200:
        current_deck = Deck.Deck(8)
        current_deck.shuffle()
    dealer_upcard: Card.Card

    # Deal
    my_cards = [current_deck.pop()]
    dealer_upcard = current_deck.pop()
    dealer_cards = [dealer_upcard]
    my_cards.append(current_deck.pop())

    if verbose: print(("My cards: " + str(my_cards)))
    if verbose: print(("Dealer cards:", dealer_cards))

    value = calc_value(my_cards)
    if verbose: print((value))

    if value == 21:  # BlackJack
        if verbose: print(("Black Jack!"))
        remember_round(value, dealer_upcard.value, 1, bj_bs_hit_table)
        remember_round(value, dealer_upcard.value, 2, bj_bs_stay_table)
        return

    while calc_value(my_cards) != 21 and not stay:
        before = calc_value(my_cards)
        my_cards.append(current_deck.pop())
        after = calc_value(my_cards)
        if verbose: print((after))
        if after > 21:  # Busted
            if verbose: print(("Bust"))
            remember_round(before, dealer_upcard.value, 1, bj_bs_hit_table)
            return
        else:  # Not Busted
            remember_round(before, dealer_upcard.value, 2, bj_bs_hit_table)
            continue

    value = calc_value(my_cards)
    if verbose: print(("Standing on " + str(value)))
    # Dealer gets cards
    if verbose: print(("Dealer has " + str(dealer_upcard.value)))
    while calc_value(dealer_cards) < 17:
        new_card = current_deck.pop()
        if verbose: print(("Dealer pulled " + str(new_card.value)))
        dealer_cards.append(new_card)
        if verbose: print(("Dealer has " + str(calc_value(dealer_cards))))

    if calc_value(dealer_cards) > 21:  # Dealer busted
        if verbose: print(("Dealer busted!"))
        dealer_cards = [Card.Card(0, False)]

    if calc_value(dealer_cards) > value:  # Loss
        remember_round(value, dealer_upcard.value, 1, bj_bs_stay_table)
        if verbose: print(("Loss"))

    elif calc_value(dealer_cards) < value:  # Win
        remember_round(value, dealer_upcard.value, 2, bj_bs_stay_table)
        if verbose: print(("Win"))
    else:  # Push
        remember_round(value, dealer_upcard.value, 0, bj_bs_stay_table)
        if verbose: print(("Push"))


def trial_and_error():
    global value
    for i in range(10):
        trial_error_game()

        if verbose: print((bj_stat))
        if verbose: print((bj_remember))
    won = 0
    lost = 0
    for value in bj_stat:
        for entry in bj_stat[value].values():
            won = won + entry[1]
            lost = lost + entry[2]
    if verbose: print(("Games won:", str(won)))
    if verbose: print(("Games lost:", str(lost)))


if __name__ == '__main__':
    # trial_and_error()
    current_deck = Deck.Deck(8)  # TODO: Einzelne Decks shufflen
    current_deck.shuffle()
    for i in range(1000000):
        stat_game(stay=False, current_deck=current_deck)
    for i in range(1000000):
        stat_game(stay=True, current_deck=current_deck)

    won = 0
    lost = 0
    for value in bj_bs_hit_table:
        for entry in bj_bs_hit_table[value].values():
            won = won + entry[1]
            lost = lost + entry[2]
    if verbose: print(("Games won with hitting:", str(won)))
    if verbose: print(("Games lost with hitting:", str(lost)))

    for value in bj_bs_stay_table:
        for entry in bj_bs_stay_table[value].values():
            won = won + entry[1]
            lost = lost + entry[2]
    if verbose: print(("Games won with staying:", str(won)))
    if verbose: print(("Games lost with staying:", str(lost)))

    if verbose: print(("Writing to excel"))

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Black Jack Basic Strategy"
    for player_upcard in range(1, 22):
        if player_upcard not in bj_bs_hit_table or player_upcard not in bj_bs_stay_table:
            continue
        for dealer_upcard in range(1, 22):
            if dealer_upcard not in bj_bs_hit_table[player_upcard] \
                    or dealer_upcard not in bj_bs_stay_table[player_upcard]:
                continue
            if verbose: print(
                ("Writing statistic for player up card", player_upcard, "with dealer up card", dealer_upcard))
            moves_played_hit = bj_bs_hit_table[player_upcard][dealer_upcard][0] or 0
            moves_won_hit = bj_bs_hit_table[player_upcard][dealer_upcard][1] or 1
            moves_lost_hit = bj_bs_hit_table[player_upcard][dealer_upcard][2] or 1
            moves_played_stay = bj_bs_stay_table[player_upcard][dealer_upcard][0] or 0
            moves_won_stay = bj_bs_stay_table[player_upcard][dealer_upcard][1] or 1
            moves_lost_stay = bj_bs_stay_table[player_upcard][dealer_upcard][2] or 1

            ratio_hit = moves_won_hit / moves_played_hit
            ratio_stay = moves_won_stay / moves_played_stay

            print("Win Ratio for Player Card", player_upcard, "and Dealer UpCard", dealer_upcard)
            print("Ratio Hit:", ratio_hit)
            print("Ratio Stay:", ratio_stay)


            if ratio_hit > ratio_stay:
                char = "H"
            elif ratio_stay > ratio_hit:
                char = "S"
            else:
                char = "More tries!"

            sheet.cell(row=player_upcard + 1, column=dealer_upcard + 1).value = char

    # Mark columns
    for i in range(2, 23):
        sheet.cell(1, i).value = i - 1
        sheet.cell(i, 1).value = i - 1

    try:
        wb.save("basic_strategy.xlsx")
        if verbose: print((bj_bs_hit_table))
        if verbose: print((bj_bs_stay_table))
    except PermissionError:
        if verbose: print(("Error Saving"))
